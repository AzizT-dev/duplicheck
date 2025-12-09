# -*- coding: utf-8 -*-
"""
DupliCheck - Result Exporter
=================================

Exports detection results to various formats:
- CSV: Simple tabular export
- Excel (XLSX): Formatted report with header and duplicate list
- GeoPackage: Spatial export with geometry preservation
"""

import os
import csv
from datetime import datetime
from typing import List, Dict, Optional

from qgis.PyQt.QtWidgets import QFileDialog, QMessageBox
from qgis.PyQt.QtCore import QCoreApplication
from qgis.core import (
    QgsVectorLayer, QgsVectorFileWriter, QgsFeature,
    QgsFeatureRequest, QgsProject, QgsField, QgsFields
)
from qgis.PyQt.QtCore import QVariant

from .detector import DuplicateGroup


class ResultExporter:
    """
    Exports duplicate detection results to various formats.
    
    Supports:
    - CSV: Lightweight tabular export
    - Excel (XLSX): Formatted report with statistics
    - GeoPackage: Full spatial export with geometries
    """
    
    def __init__(
        self,
        groups: List[DuplicateGroup],
        layer: QgsVectorLayer,
        actions: Dict[int, str] = None,
        config: Dict = None
    ):
        """
        Initialize the exporter.
        
        :param groups: List of DuplicateGroup objects
        :param layer: Source vector layer
        :param actions: Dictionary mapping feature IDs to actions ('keep'/'remove')
        :param config: Detection configuration dictionary
        """
        self.groups = groups
        self.layer = layer
        self.actions = actions or {}
        self.config = config or {}
    
    def tr(self, message):
        """Get translation for a string."""
        return QCoreApplication.translate('ResultExporter', message)
    
    def export(self, format_type: str, parent=None) -> Optional[str]:
        """
        Export results to the specified format.
        
        :param format_type: Export format ('csv', 'xlsx', 'gpkg')
        :param parent: Parent widget for dialogs
        :returns: Output file path or None if cancelled
        """
        # Get output path from user
        filters = {
            'csv': self.tr('CSV Files (*.csv)'),
            'xlsx': self.tr('Excel Files (*.xlsx)'),
            'gpkg': self.tr('GeoPackage Files (*.gpkg)')
        }
        
        default_name = f"{self.layer.name()}_duplicates_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            self.tr('Export Results'),
            default_name,
            filters.get(format_type, filters['csv'])
        )
        
        if not file_path:
            return None
        
        # Ensure correct extension
        extensions = {'csv': '.csv', 'xlsx': '.xlsx', 'gpkg': '.gpkg'}
        ext = extensions.get(format_type, '.csv')
        if not file_path.lower().endswith(ext):
            file_path += ext
        
        # Export based on format
        if format_type == 'csv':
            return self._export_csv(file_path)
        elif format_type == 'xlsx':
            return self._export_xlsx(file_path)
        elif format_type == 'gpkg':
            return self._export_gpkg(file_path)
        else:
            raise ValueError(f"Unknown format: {format_type}")
    
    def _get_id_field(self):
        """Get the ID field name from config."""
        return self.config.get('id_field') or (
            self.layer.fields()[0].name() if self.layer.fields().count() > 0 else None
        )
    
    def _get_feature_id_values(self):
        """
        Get a mapping of FID to ID field value.
        
        :returns: Dictionary {fid: id_value}
        """
        id_field = self._get_id_field()
        feature_values = {}
        
        for feature in self.layer.getFeatures():
            fid = feature.id()
            if id_field and id_field in [f.name() for f in self.layer.fields()]:
                feature_values[fid] = str(feature[id_field])
            else:
                feature_values[fid] = str(fid)
        
        return feature_values
    
    def _export_csv(self, file_path: str) -> str:
        """
        Export results to CSV format - simple list of duplicates.
        
        :param file_path: Output file path
        :returns: File path
        """
        id_field = self._get_id_field()
        feature_values = self._get_feature_id_values()
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                self.tr('N°'),
                self.tr('IDs Entités')
            ])
            
            # Data rows - one row per group
            for i, group in enumerate(self.groups):
                # Get ID values for all features in group
                id_values = [feature_values.get(fid, str(fid)) for fid in group.feature_ids]
                ids_str = ', '.join(id_values)
                
                writer.writerow([
                    i + 1,
                    ids_str
                ])
        
        return file_path
    
    def _export_xlsx(self, file_path: str) -> str:
        """
        Export results to Excel format - header with summary + simple list.
        
        Format requested by user:
        - Header section with report info
        - Simple table: N° | IDs Entités
        
        :param file_path: Output file path
        :returns: File path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not available
            csv_path = file_path.replace('.xlsx', '.csv')
            return self._export_csv(csv_path)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.tr('Duplicates Report')
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        label_font = Font(bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === HEADER SECTION ===
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value='DupliCheck - Detection Report')
        ws.cell(row=row, column=1).font = title_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        # Report info
        ws.cell(row=row, column=1, value=self.tr('Généré le :')).font = label_font
        ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 1
        
        ws.cell(row=row, column=1, value=self.tr('Couche :')).font = label_font
        ws.cell(row=row, column=2, value=self.layer.name())
        row += 1
        
        # Detection type
        detection_type = self.config.get('detection_type', 'geometry')
        type_label = self.tr('Géométrique') if detection_type == 'geometry' else self.tr('Attributaire')
        ws.cell(row=row, column=1, value=self.tr('Type :')).font = label_font
        ws.cell(row=row, column=2, value=type_label)
        row += 1
        
        ws.cell(row=row, column=1, value=self.tr('Entités Totales :')).font = label_font
        ws.cell(row=row, column=2, value=self.layer.featureCount())
        row += 2
        
        # Statistics section
        ws.cell(row=row, column=1, value=self.tr('Statistiques de Détection'))
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        total_groups = len(self.groups)
        total_duplicates = sum(len(g.feature_ids) for g in self.groups)
        
        ws.cell(row=row, column=1, value=self.tr('Groupes de Doublons :')).font = label_font
        ws.cell(row=row, column=2, value=total_groups)
        row += 1
        
        ws.cell(row=row, column=1, value=self.tr('Entités en Doublon :')).font = label_font
        ws.cell(row=row, column=2, value=total_duplicates)
        row += 2
        
        # === DATA TABLE ===
        # Table header
        table_start_row = row
        
        headers = [self.tr('N°'), self.tr('IDs Entités')]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1
        
        # Get feature ID values
        feature_values = self._get_feature_id_values()
        
        # Data rows
        for i, group in enumerate(self.groups):
            # Get ID values for all features in group
            id_values = [feature_values.get(fid, str(fid)) for fid in group.feature_ids]
            ids_str = ', '.join(id_values)
            
            # N° column
            cell = ws.cell(row=row, column=1, value=i + 1)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # IDs column
            cell = ws.cell(row=row, column=2, value=ids_str)
            cell.border = thin_border
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
        
        # Save
        wb.save(file_path)
        return file_path
    
    def _export_gpkg(self, file_path: str) -> str:
        """
        Export results to GeoPackage with full geometry.
        
        Creates a layer with all duplicate features and their metadata.
        
        :param file_path: Output file path
        :returns: File path
        """
        # Get all duplicate feature IDs
        all_fids = set()
        fid_to_group = {}
        
        for i, group in enumerate(self.groups):
            for fid in group.feature_ids:
                all_fids.add(fid)
                fid_to_group[fid] = (i + 1, group)
        
        # Create fields
        fields = QgsFields()
        
        # Copy original fields
        for field in self.layer.fields():
            fields.append(field)
        
        # Add duplicate metadata fields
        fields.append(QgsField('dup_group_id', QVariant.Int))
        fields.append(QgsField('dup_type', QVariant.String))
        fields.append(QgsField('dup_action', QVariant.String))
        
        # Get features
        request = QgsFeatureRequest().setFilterFids(list(all_fids))
        
        # Write to GeoPackage
        options = QgsVectorFileWriter.SaveVectorOptions()
        options.driverName = 'GPKG'
        options.fileEncoding = 'UTF-8'
        
        # Determine geometry type
        geom_type = self.layer.wkbType()
        crs = self.layer.crs()
        
        writer = QgsVectorFileWriter.create(
            file_path,
            fields,
            geom_type,
            crs,
            QgsProject.instance().transformContext(),
            options
        )
        
        if writer.hasError() != QgsVectorFileWriter.NoError:
            raise IOError(f"Error creating GeoPackage: {writer.errorMessage()}")
        
        # Write features
        for feature in self.layer.getFeatures(request):
            fid = feature.id()
            group_id, group = fid_to_group.get(fid, (0, None))
            
            # Create new feature
            new_feature = QgsFeature(fields)
            new_feature.setGeometry(feature.geometry())
            
            # Copy original attributes
            for i, field in enumerate(self.layer.fields()):
                new_feature[field.name()] = feature[field.name()]
            
            # Add metadata
            if group:
                new_feature['dup_group_id'] = group_id
                new_feature['dup_type'] = group.detection_type
                new_feature['dup_action'] = self.actions.get(fid, '')
            
            writer.addFeature(new_feature)
        
        del writer
        return file_path


class SnapshotManager:
    """
    Manages layer snapshots for undo/restore functionality.
    """
    
    def __init__(self, temp_dir: str = None):
        """
        Initialize the snapshot manager.
        
        :param temp_dir: Directory for temporary snapshots
        """
        import tempfile
        self.temp_dir = temp_dir or tempfile.gettempdir()
        self.snapshots = {}  # layer_id -> snapshot_path
    
    def create_snapshot(self, layer: QgsVectorLayer) -> str:
        """
        Create a snapshot of a layer.
        
        :param layer: Layer to snapshot
        :returns: Snapshot file path
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        snapshot_path = os.path.join(
            self.temp_dir,
            f"duplicheck_snapshot_{layer.id()}_{timestamp}.gpkg"
        )
        
        options = QgsVectorFileWriter.SaveVectorOptions()
        options.driverName = 'GPKG'
        options.fileEncoding = 'UTF-8'
        
        QgsVectorFileWriter.writeAsVectorFormatV3(
            layer,
            snapshot_path,
            QgsProject.instance().transformContext(),
            options
        )
        
        self.snapshots[layer.id()] = snapshot_path
        return snapshot_path
    
    def restore_snapshot(self, layer_id: str) -> Optional[QgsVectorLayer]:
        """
        Restore a layer from snapshot.
        
        :param layer_id: ID of the layer to restore
        :returns: Restored layer or None
        """
        snapshot_path = self.snapshots.get(layer_id)
        
        if not snapshot_path or not os.path.exists(snapshot_path):
            return None
        
        restored = QgsVectorLayer(snapshot_path, "Restored Layer", "ogr")
        
        if restored.isValid():
            return restored
        
        return None
    
    def cleanup(self):
        """Remove all temporary snapshot files."""
        for path in self.snapshots.values():
            try:
                if os.path.exists(path):
                    os.remove(path)
            except:
                pass
        
        self.snapshots.clear()
